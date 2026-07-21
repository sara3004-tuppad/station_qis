"""
Read/write the Stations Movement Tracker Excel on SharePoint.

In dev_mode the local cache file is used directly (no network calls).
In production the file is fetched/saved via Power Automate HTTP flows.

Caching strategy:
- read_quality_sheet / read_allocation_sheet are cached (ttl=300s fallback)
- every write calls st.cache_data.clear() so the next read is always fresh
"""

import io
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from utils.config import get_config


def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    """Coerce all object-dtype columns to strings so PyArrow can serialize them.
    Object columns are the source of mixed int64/str/bytes type errors."""
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str).replace("nan", "")
    return df


def _is_dev() -> bool:
    return get_config().get("dev_mode", False)


def _local_path() -> Path:
    cfg = get_config()
    return Path(__file__).parent.parent / cfg["excel"]["local_cache_path"]


# ---------------------------------------------------------------------------
# Core: get workbook bytes (from SharePoint or local)
# ---------------------------------------------------------------------------

def _download_wb_bytes() -> bytes:
    if _is_dev():
        return _local_path().read_bytes()
    from utils.graph_client import download_excel_from_sharepoint
    return download_excel_from_sharepoint(get_config()["excel"]["sharepoint_url"])


def _upload_wb_bytes(data: bytes):
    if _is_dev():
        _local_path().write_bytes(data)
        return
    from utils.graph_client import upload_excel_to_sharepoint
    upload_excel_to_sharepoint(get_config()["excel"]["sharepoint_url"], data)


@contextmanager
def _open_workbook():
    """Download, yield an openpyxl workbook, upload saved bytes, then clear cache."""
    raw = _download_wb_bytes()
    wb = load_workbook(io.BytesIO(raw))
    yield wb
    buf = io.BytesIO()
    wb.save(buf)
    _upload_wb_bytes(buf.getvalue())
    st.cache_data.clear()


@st.cache_data(ttl=300)
def _cached_wb_bytes() -> bytes:
    return _download_wb_bytes()


def _read_sheet_df(sheet_name: str) -> pd.DataFrame:
    raw = _cached_wb_bytes()
    df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet_name, header=1)
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    return df


# ---------------------------------------------------------------------------
# Quality Clearance Sheet
# ---------------------------------------------------------------------------

@st.cache_data(ttl=300)
def read_quality_sheet() -> pd.DataFrame:
    cfg = get_config()
    df = _read_sheet_df(cfg["excel"]["quality_sheet"])
    return _sanitize_df(df.dropna(how="all"))


def append_quality_row(row: dict):
    cfg = get_config()
    sheet_name = cfg["excel"]["quality_sheet"]
    with _open_workbook() as wb:
        ws = wb[sheet_name]
        for r in range(3, ws.max_row + 2):
            if ws.cell(row=r, column=2).value is None:
                first_empty = r
                break
        ws.cell(row=first_empty, column=1, value=first_empty - 2)
        ws.cell(row=first_empty, column=2, value=row.get("Date"))
        ws.cell(row=first_empty, column=3, value=row.get("Date").strftime("%B") if row.get("Date") else "")
        ws.cell(row=first_empty, column=4, value=row.get("QIS No"))
        ws.cell(row=first_empty, column=5, value=row.get("Type"))
        ws.cell(row=first_empty, column=6, value=row.get("PDI"))


def update_quality_row_stores(qis_no: str, pickup_status: str, grn_date):
    cfg = get_config()
    sheet_name = cfg["excel"]["quality_sheet"]
    with _open_workbook() as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3):
            if str(row[3].value) == str(qis_no):
                row[6].value = pickup_status
                row[7].value = grn_date
                break


def bulk_update_quality_rows_stores(updates: dict):
    """Write pickup status and GRN date for multiple QIS numbers in one workbook open.
    updates: {qis_no: {"pickup_status": str, "grn_date": date or None}}
    """
    cfg = get_config()
    sheet_name = cfg["excel"]["quality_sheet"]
    with _open_workbook() as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3):
            qis_no = str(row[3].value)
            if qis_no in updates:
                entry = updates[qis_no]
                row[6].value = entry.get("pickup_status") or ""
                row[7].value = entry.get("grn_date")


# ---------------------------------------------------------------------------
# Allocation Sheet
# ---------------------------------------------------------------------------

@st.cache_data(ttl=300)
def read_allocation_sheet() -> pd.DataFrame:
    cfg = get_config()
    df = _read_sheet_df(cfg["excel"]["allocation_sheet"])
    return _sanitize_df(df.dropna(subset=["Site Name"]))


def update_cpt_row(site_name: str, allocation: str, remarks: str):
    cfg = get_config()
    sheet_name = cfg["excel"]["allocation_sheet"]
    with _open_workbook() as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3):
            if row[1].value == site_name:
                row[12].value = allocation
                row[13].value = remarks
                break


def update_stores_dispatch_row(site_name: str, data: dict):
    cfg = get_config()
    sheet_name = cfg["excel"]["allocation_sheet"]
    with _open_workbook() as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3):
            if row[1].value == site_name:
                row[14].value = data.get("Date of Dispatch")
                row[15].value = data.get("No of QIS delivered")
                row[16].value = data.get("Expected date of delivery")
                row[17].value = data.get("E-Way Bill")
                break


def mark_email_sent(site_name: str):
    cfg = get_config()
    sheet_name = cfg["excel"]["allocation_sheet"]
    with _open_workbook() as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3):
            if row[1].value == site_name:
                row[18].value = "Yes"
                row[19].value = "Sent"
                row[20].value = datetime.now().date()
                break


# ---------------------------------------------------------------------------
# Completion checks — use cached reads, no extra download
# ---------------------------------------------------------------------------

def is_quality_row_complete(qis_no: str) -> bool:
    cfg = get_config()
    fields = cfg["email"]["completion_check"]["quality_clearance"]
    col_map = {
        "Date": "Date", "QIS No": "QIS No", "Type": "Type",
        "PDI": "PDI", "Pickup status": "Pickup status", "GRN Date": "GRN Date",
    }
    df = read_quality_sheet()
    rows = df[df["QIS No"].astype(str) == str(qis_no)]
    if rows.empty:
        return False
    row = rows.iloc[0]
    return all(pd.notna(row.get(col_map[f])) and str(row.get(col_map[f])).strip() != "" for f in fields if f in col_map)


def is_allocation_row_complete(site_name: str) -> bool:
    cfg = get_config()
    fields = cfg["email"]["completion_check"]["allocation"]
    col_map = {
        "Allocation": "Allocation", "Date of Dispatch": "Date of Dispatch",
        "No of QIS delivered": "No of QIS delivered", "E-Way Bill": "E-Way Bill",
    }
    df = read_allocation_sheet()
    rows = df[df["Site Name"] == site_name]
    if rows.empty:
        return False
    row = rows.iloc[0]
    return all(pd.notna(row.get(col_map[f])) and str(row.get(col_map[f])).strip() != "" for f in fields if f in col_map)
